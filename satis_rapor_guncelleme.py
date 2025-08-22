
# -*- coding: utf-8 -*-
import json, time
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import numpy as np
from openpyxl import load_workbook

APP_NAME = "satis_rapor_guncelleme"
CONFIG_PATH = Path.home() / f".{APP_NAME}.json"
MONTH_SHEETS = ["OCAK","ŞUBAT","MART","NİSAN","MAYIS","HAZİRAN","TEMMUZ","AĞUSTOS","EYLÜL","EKİM","KASIM","ARALIK"]

def parse_turkish_number(val):
    if pd.isna(val): return 0.0
    s = str(val).strip().replace(" ","").replace(" ","").replace(".","").replace(",",".")
    try: return float(s)
    except: return 0.0

def to_num(series): return series.apply(parse_turkish_number)
def nrm(s): return str(s).strip().upper() if s is not None else ""

def map_birim_to_report(b):
    b = nrm(b)
    if "TOSYA" in b: return "TOSYA"
    if "ANADOLU" in b: return "BOYA-2"
    if "AVRUPA" in b: return "BOYA"
    if "DEPAR" in b: return "DEPARKO"
    return b

CAND = {
    "tarih": ["Tarih","Gün","Gun","Date"],
    "birim": ["Birim","Birim ","Birim ","Birim  ","Birim Adı","Unit","Ünite","SATIŞ YERİ","SATIS YERI"],
    "toptan": ["Toptan","Toptan Satış","Toptan Satis","Depo Satış","Depo Satis","TOPTAN"],
    "fabrika": ["Fabrika","Fabrika Satış","Fabrika Satis","FABRİKADAN"],
    "ihracat": ["İhracat","İhracat Satış","Ihracat Satis","Export"],
    "nakit": ["Nakit","NAKİT"],
    "havale_toplam": ["Havale","HAVALE"],
    "cek": ["Çek","CEK","ÇEK"],
    "pos_kendi": ["Kendi POS","Kendi Posumuz","KENDİ POSUMUZ","POS Kendi"],
    "akbank": ["Akbank","AKBANK"],
    "isbank": ["İşbank","İş Bankası","İŞ BANKASI","Isbank"],
    "garanti": ["Garanti","GARANTİ"],
    "vakifbank": ["Vakıfbank","VAKIFBANK","Vakıf Bank"],
    "pos_kastamonu": ["Kastamonu Entegre","KASTAMONU ENTEGRE"],
    "pos_kayalar": ["Kayalar","KAYALAR"],
    "pos_camsar": ["Çamsar","ÇAMSAR","CAMSAR"],
    "pos_sfc": ["SFC Entegre","SFC ENTREGRE","SFC"],
    "pos_camsan": ["Çamsan Ordu","ÇAMSAN ORDU","CAMSAN ORDU"],
}

def find_col(df, keys):
    norm = {str(c).strip().lower(): c for c in df.columns}
    for key in keys:
        k = key.strip().lower()
        if k in norm: return norm[k]
        for nc, orig in norm.items():
            if k in nc: return orig
    return None

def load_sales(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path, dtype=object)
    cols = {k: find_col(df, v) for k, v in CAND.items()}
    out = pd.DataFrame()
    out["Tarih"] = pd.to_datetime(df[cols["tarih"]], errors="coerce")
    out["Birim"] = df[cols["birim"]].astype(str).map(nrm)
    def numcol(k):
        c = cols.get(k)
        return to_num(df[c]) if c is not None else pd.Series(0, index=df.index)
    out["Toptan"] = numcol("toptan"); out["Fabrika"] = numcol("fabrika"); out["İhracat"] = numcol("ihracat")
    out["Nakit"] = numcol("nakit")
    out["Havale_Akbank"] = numcol("akbank"); out["Havale_İşbank"] = numcol("isbank")
    out["Havale_Garanti"] = numcol("garanti"); out["Havale_Vakıfbank"] = numcol("vakifbank")
    hv_banks = out[["Havale_Akbank","Havale_İşbank","Havale_Garanti","Havale_Vakıfbank"]].sum(axis=1)
    hv_total = numcol("havale_toplam"); out["Havale"] = np.where(hv_total>0, hv_total, hv_banks)
    out["Çek"] = numcol("cek")
    out["POS_Kendi"] = numcol("pos_kendi")
    out["POS_Kastamonu"] = numcol("pos_kastamonu"); out["POS_Kayalar"] = numcol("pos_kayalar")
    out["POS_Çamsar"] = numcol("pos_camsar"); out["POS_SFC"] = numcol("pos_sfc"); out["POS_Çamsan"] = numcol("pos_camsan")
    out["POS_Toplam"] = out["POS_Kendi"] + out[["POS_Kastamonu","POS_Kayalar","POS_Çamsar","POS_SFC","POS_Çamsan"]].sum(axis=1)
    out = out.dropna(subset=["Tarih"])
    val_cols = ["Toptan","Fabrika","İhracat","Nakit","Havale","Çek","POS_Kendi","Havale_Akbank","Havale_İşbank","Havale_Garanti","Havale_Vakıfbank","POS_Kastamonu","POS_Kayalar","POS_Çamsar","POS_SFC","POS_Çamsan","POS_Toplam"]
    grouped = (out.groupby(["Birim","Tarih"], as_index=False)[val_cols].sum().sort_values(["Birim","Tarih"]).reset_index(drop=True))
    grouped["SatışToplam"] = grouped[["Toptan","Fabrika","İhracat"]].sum(axis=1)
    grouped["TahsilatToplam"] = grouped[["Nakit","Havale","Çek","POS_Toplam"]].sum(axis=1)
    grouped["_BIRIM_REPORT"] = grouped["Birim"].map(map_birim_to_report)
    return grouped

FIELD_TO_TARGETS = {
    "Toptan": [("SATIŞ", "TOPTAN SATIŞ"), ("", "TOPTAN SATIŞ")],
    "Fabrika": [("SATIŞ", "FABRİKADAN SATIŞ"), ("", "FABRİKADAN SATIŞ")],
    "İhracat": [("SATIŞ", "İHRACAT SATIŞ"), ("", "İHRACAT SATIŞ")],
    "SatışToplam": [("SATIŞ", "SATIŞ TOPLAMI"), ("", "SATIŞ TOPLAMI")],
    "POS_Kendi": [("KREDİ KARTI POS","KENDİ POSUMUZ")],
    "POS_Kastamonu": [("KREDİ KARTI POS","KASTAMONU ENTEGRE")],
    "POS_Kayalar": [("KREDİ KARTI POS","KAYALAR")],
    "POS_Çamsar": [("KREDİ KARTI POS","ÇAMSAR")],
    "POS_SFC": [("KREDİ KARTI POS","SFC ENTEGRE")],
    "POS_Çamsan": [("KREDİ KARTI POS","ÇAMSAN ORDU")],
    "Nakit":  [("", "NAKİT")],
    "Havale": [("", "HAVALE")],
    "Çek":    [("", "ÇEK")],
    "TahsilatToplam": [("", "TAHSİLAT TOPLAMI")],
    "Havale_Akbank":  [("GELEN HAVALE","AKBANK")],
    "Havale_İşbank":  [("GELEN HAVALE","İŞ BANKASI")],
    "Havale_Garanti": [("GELEN HAVALE","GARANTİ")],
    "Havale_Vakıfbank":[("GELEN HAVALE","VAKIFBANK")],
}

def build_block_map(ws):
    blk = [ws.cell(1,c).value for c in range(1, ws.max_column+1)]
    sub = [ws.cell(2,c).value for c in range(1, ws.max_column+1)]
    m = {}
    for i,(b,s) in enumerate(zip(blk, sub)):
        bb = nrm(b) if b is not None else ""
        ss = nrm(s) if s is not None else ""
        m[(bb, ss)] = i+1
    return m, sub

def find_col_by_sub_only(sub_headers, sub_name):
    sub_name = nrm(sub_name)
    for i, s in enumerate(sub_headers):
        if nrm(s) == sub_name:
            return i+1
    return None

def target_col_for_field(field, block_map, sub_headers):
    for blk, sub in FIELD_TO_TARGETS[field]:
        col = block_map.get((nrm(blk), nrm(sub)))
        if col: return col
    for blk, sub in FIELD_TO_TARGETS[field]:
        col = find_col_by_sub_only(sub_headers, sub)
        if col: return col
    return None

def update_report(sales_path: Path, report_path: Path, save_mode: str, log=lambda *a, **k: None) -> Path:
    log("Satış verisi yükleniyor…")
    sales_all = load_sales(sales_path)
    log("Rapor dosyası açılıyor…")
    wb = load_workbook(report_path)
    total_updates = []
    for month_idx, sheet in enumerate(MONTH_SHEETS, start=1):
        if sheet not in wb.sheetnames: continue
        ws = wb[sheet]
        block_map, sub_headers = build_block_map(ws)
        df_sheet = pd.read_excel(report_path, sheet_name=sheet, header=1)
        if "SATIŞ YERİ" not in df_sheet.columns or "GÜN" not in df_sheet.columns:
            total_updates.append((sheet, 0, "SATIŞ YERİ/GÜN yok")); continue
        df_sheet["SATIŞ YERİ"] = df_sheet["SATIŞ YERİ"].ffill()
        df_sheet["GÜN"] = df_sheet["GÜN"].ffill()
        df_sheet["_EXCEL_ROW"] = df_sheet.index + 3
        df_sheet["_BIRIM_NORM"] = df_sheet["SATIŞ YERİ"].map(nrm)
        df_sheet["_GUN_DATE"] = pd.to_datetime(df_sheet["GÜN"], errors="coerce", dayfirst=True)
        sales_m = sales_all[(sales_all["Tarih"].dt.year==2025) & (sales_all["Tarih"].dt.month==month_idx)].copy()
        if sales_m.empty:
            total_updates.append((sheet, 0, "Satış yok")); continue
        updated_rows = 0
        for _, srow in sales_m.iterrows():
            gun = pd.to_datetime(srow["Tarih"]).date()
            birim_key = srow["_BIRIM_REPORT"]
            candidates = df_sheet[(df_sheet["_GUN_DATE"].dt.date == gun) & (df_sheet["_BIRIM_NORM"] == nrm(birim_key))]
            if candidates.empty: continue
            excel_row = int(candidates.iloc[0]["_EXCEL_ROW"])
            for field in ["Toptan","Fabrika","İhracat","SatışToplam","POS_Kendi","POS_Kastamonu","POS_Kayalar","POS_Çamsar","POS_SFC","POS_Çamsan","Nakit","Havale","Çek","TahsilatToplam","Havale_Akbank","Havale_İşbank","Havale_Garanti","Havale_Vakıfbank"]:
                col = target_col_for_field(field, block_map, sub_headers)
                if not col: continue
                val = srow.get(field, None)
                if val is None: continue
                try: ws.cell(excel_row, col).value = float(val)
                except: ws.cell(excel_row, col).value = val
            updated_rows += 1
        total_updates.append((sheet, updated_rows, "OK"))
    if save_mode == "copy":
        ts = time.strftime("%Y%m%d_%H%M%S")
        out_path = report_path.with_name(report_path.stem + f"_GUNCEL_{ts}" + report_path.suffix)
        wb.save(out_path); log(f"Kayıt (kopya): {out_path}"); return out_path
    else:
        wb.save(report_path); log(f"Kayıt (yerinde): {report_path}"); return report_path

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Satış Rapor Güncelleme"); self.geometry("780x520"); self.minsize(720,480)
        self.sales_path = tk.StringVar(); self.report_path = tk.StringVar(); self.save_mode = tk.StringVar(value="copy")
        self._load_config()
        frm_top = ttk.Frame(self); frm_top.pack(fill="x", padx=12, pady=8)
        ttk.Label(frm_top, text="Satışlar dosyası (satışlar.xlsx):").grid(row=0,column=0,sticky="w")
        ttk.Entry(frm_top, textvariable=self.sales_path, width=70).grid(row=0,column=1,padx=6)
        ttk.Button(frm_top, text="Seç…", command=self.pick_sales).grid(row=0,column=2)
        ttk.Label(frm_top, text="Rapor dosyası (SATIŞ RAPORLARI 2025.xlsx):").grid(row=1,column=0,sticky="w",pady=(6,0))
        ttk.Entry(frm_top, textvariable=self.report_path, width=70).grid(row=1,column=1,padx=6,pady=(6,0))
        ttk.Button(frm_top, text="Seç…", command=self.pick_report).grid(row=1,column=2,pady=(6,0))
        frm_mode = ttk.Frame(self); frm_mode.pack(fill="x", padx=12, pady=6)
        ttk.Label(frm_mode, text="Kayıt yöntemi:").pack(side="left")
        ttk.Radiobutton(frm_mode, text="Kopya oluştur (önerilen)", variable=self.save_mode, value="copy").pack(side="left", padx=8)
        ttk.Radiobutton(frm_mode, text="Yerinde güncelle", variable=self.save_mode, value="inplace").pack(side="left", padx=8)
        ttk.Button(frm_mode, text="Ayarları Kaydet", command=self.save_config).pack(side="right")
        frm_btn = ttk.Frame(self); frm_btn.pack(fill="x", padx=12, pady=6)
        ttk.Button(frm_btn, text="Güncellemeyi Çalıştır", command=self.run_update).pack(side="left")
        ttk.Button(frm_btn, text="Çıkış", command=self.destroy).pack(side="right")
        frm_log = ttk.LabelFrame(self, text="Kayıt / Log"); frm_log.pack(fill="both", expand=True, padx=12, pady=8)
        self.txt = tk.Text(frm_log, height=18); self.txt.pack(fill="both", expand=True)
        if not self.sales_path.get() or not self.report_path.get():
            self.log("Lütfen dosya konumlarını seçin ve 'Ayarları Kaydet' deyin.")
    def log(self, msg): self.txt.insert("end", msg+"\n"); self.txt.see("end"); self.update_idletasks()
    def pick_sales(self):
        p = filedialog.askopenfilename(title="satışlar.xlsx seçin", filetypes=[("Excel","*.xlsx *.xls")])
        if p: self.sales_path.set(p)
    def pick_report(self):
        p = filedialog.askopenfilename(title="SATIŞ RAPORLARI 2025.xlsx seçin", filetypes=[("Excel","*.xlsx *.xls")])
        if p: self.report_path.set(p)
    def _load_config(self):
        if CONFIG_PATH.exists():
            try:
                data = json.load(open(CONFIG_PATH,"r",encoding="utf-8"))
                self.sales_path.set(data.get("sales_path",""))
                self.report_path.set(data.get("report_path",""))
                self.save_mode.set(data.get("save_mode","copy"))
            except Exception as e: print("Ayar okunamadı:", e)
    def save_config(self):
        d = {"sales_path": self.sales_path.get(),"report_path": self.report_path.get(),"save_mode": self.save_mode.get()}
        try:
            json.dump(d, open(CONFIG_PATH,"w",encoding="utf-8"), ensure_ascii=False, indent=2)
            self.log(f"Ayarlar kaydedildi: {CONFIG_PATH}")
        except Exception as e:
            messagebox.showerror("Hata", f"Ayar kaydedilemedi: {e}")
    def run_update(self):
        sp = Path(self.sales_path.get()); rp = Path(self.report_path.get())
        if not sp.exists() or not rp.exists():
            messagebox.showwarning("Eksik", "Dosya konumları geçersiz. Lütfen kontrol edin."); return
        self.save_config()
        try:
            out_path = update_report(sp, rp, self.save_mode.get(), log=self.log)
            self.log("Tamamlandı."); messagebox.showinfo("Bitti", f"Güncelleme tamamlandı.\nÇıktı: {out_path}")
        except Exception as e:
            self.log(f"HATA: {e}"); messagebox.showerror("Hata", str(e))
if __name__ == "__main__":
    App().mainloop()
